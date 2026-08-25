import csv
import io
import json
import os
import re
import shutil
import sqlite3
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Body
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
import subprocess
import tempfile
import urllib.parse
import urllib.request
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "uploads"
CHUNK_DIR = BASE_DIR / "chunks"
EXPORT_DIR = BASE_DIR / "exports"
DB_PATH = DATA_DIR / "app.db"
APP_PORT = int(os.environ.get("PORT", "8099"))
for p in (DATA_DIR, UPLOAD_DIR, CHUNK_DIR, EXPORT_DIR):
    p.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="THEO DÕI CASA KHÁCH HÀNG")
app.mount("/static", StaticFiles(directory=str(PUBLIC_DIR)), name="static")

TNR_DIR = Path.home() / ".local" / "share" / "fonts" / "msttcorefonts"
pdfmetrics.registerFont(TTFont("TimesNewRoman", str(TNR_DIR / "times.ttf")))
pdfmetrics.registerFont(TTFont("TimesNewRoman-Bold", str(TNR_DIR / "timesbd.ttf")))


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def conn():
    c = sqlite3.connect(DB_PATH, timeout=120)
    c.execute("PRAGMA busy_timeout=120000")
    c.row_factory = sqlite3.Row
    return c


def load_dkdv_tktt_targets(limit: int = 5000) -> List[Dict[str, str]]:
    """Read TKTT hoạt động registrations from dkdv app (8091)."""
    url = f"http://127.0.0.1:8091/api/registrations?limit={int(limit)}"
    try:
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    rows = data.get("records") or data.get("rows") or []
    out: List[Dict[str, str]] = []
    seen = set()
    for r in rows:
        fields = r.get("fields") if isinstance(r, dict) else None
        if not isinstance(fields, dict):
            fields = r if isinstance(r, dict) else {}
        service = str(fields.get("Loại dịch vụ") or fields.get("service_type") or "").strip().lower()
        if "tktt" not in service or "hoạt động" not in service:
            continue
        account = re.sub(r"\D", "", str(fields.get("Số tài khoản") or fields.get("account_number") or ""))
        officer = str(fields.get("Tên cán bộ") or fields.get("staff_name") or "").strip()
        customer_name = str(fields.get("Tên khách hàng") or fields.get("customer_name") or "").strip()
        if not account or not officer:
            continue
        key = (account, officer)
        if key in seen:
            continue
        seen.add(key)
        out.append({"account_number": account, "officer_name": officer, "customer_name": customer_name})
    return out


def init_db():
    with conn() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS datasets(
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            original_name TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            file_type TEXT NOT NULL,
            columns_json TEXT NOT NULL,
            row_count INTEGER NOT NULL DEFAULT 0,
            imported_at TEXT NOT NULL,
            dataset_kind TEXT DEFAULT 'transaction'
        )
        """)
        ds_cols = [r[1] for r in c.execute("PRAGMA table_info(datasets)").fetchall()]
        if "dataset_kind" not in ds_cols:
            c.execute("ALTER TABLE datasets ADD COLUMN dataset_kind TEXT DEFAULT 'transaction'")
        c.execute("""
        CREATE TABLE IF NOT EXISTS rows(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_id TEXT NOT NULL,
            row_no INTEGER NOT NULL,
            data_json TEXT NOT NULL,
            search_text TEXT NOT NULL,
            FOREIGN KEY(dataset_id) REFERENCES datasets(id) ON DELETE CASCADE
        )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_rows_dataset ON rows(dataset_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_rows_search ON rows(search_text)")
        c.execute("""CREATE TABLE IF NOT EXISTS app_users(id TEXT PRIMARY KEY, username TEXT NOT NULL UNIQUE, full_name TEXT NOT NULL, department TEXT DEFAULT '', unit TEXT DEFAULT '', created_at TEXT NOT NULL)""")
        c.execute("""CREATE TABLE IF NOT EXISTS user_groups(id TEXT PRIMARY KEY, name TEXT NOT NULL UNIQUE, note TEXT DEFAULT '', created_at TEXT NOT NULL)""")
        c.execute("""CREATE TABLE IF NOT EXISTS user_group_members(group_id TEXT NOT NULL, user_id TEXT NOT NULL, PRIMARY KEY(group_id,user_id))""")
        c.execute("""CREATE TABLE IF NOT EXISTS customer_targets(id TEXT PRIMARY KEY, customer_code TEXT NOT NULL, customer_name TEXT DEFAULT '', officer_name TEXT NOT NULL, note TEXT DEFAULT '', created_at TEXT NOT NULL)""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_customer_targets_code ON customer_targets(customer_code)")
        c.execute("""
        CREATE TABLE IF NOT EXISTS report_rows(
            dataset_id TEXT NOT NULL,
            row_no INTEGER NOT NULL,
            tr_date TEXT DEFAULT '',
            customer_code TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            currency TEXT DEFAULT '',
            debit REAL DEFAULT 0,
            credit REAL DEFAULT 0,
            turnover REAL DEFAULT 0,
            search_text TEXT DEFAULT '',
            PRIMARY KEY(dataset_id,row_no),
            FOREIGN KEY(dataset_id) REFERENCES datasets(id) ON DELETE CASCADE
        )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_report_rows_dataset_date ON report_rows(dataset_id,tr_date)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_report_rows_date ON report_rows(tr_date)")
        cols = [r[1] for r in c.execute("PRAGMA table_info(report_rows)").fetchall()]
        if "currency" not in cols:
            c.execute("ALTER TABLE report_rows ADD COLUMN currency TEXT DEFAULT ''")
        c.execute("CREATE INDEX IF NOT EXISTS idx_report_rows_customer ON report_rows(customer_code)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_report_rows_currency ON report_rows(currency)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_report_rows_search ON report_rows(search_text)")
        c.execute("""
        CREATE TABLE IF NOT EXISTS account_balance_rows(
            dataset_id TEXT NOT NULL,
            row_no INTEGER NOT NULL,
            value_date TEXT NOT NULL,
            branch_code TEXT DEFAULT '',
            account_ledger TEXT DEFAULT '',
            customer_code TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            currency TEXT DEFAULT '',
            current_balance REAL DEFAULT 0,
            rate REAL DEFAULT 0,
            account_number TEXT DEFAULT '',
            search_text TEXT DEFAULT '',
            PRIMARY KEY(dataset_id,row_no),
            FOREIGN KEY(dataset_id) REFERENCES datasets(id) ON DELETE CASCADE
        )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_balance_rows_date ON account_balance_rows(value_date)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_balance_rows_dataset_date ON account_balance_rows(dataset_id,value_date)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_balance_rows_customer ON account_balance_rows(customer_code)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_balance_rows_account ON account_balance_rows(account_number)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_balance_rows_search ON account_balance_rows(search_text)")
        # CASA calculation source: account ledger 421101 only. Ledger 211108 is
        # retained separately only for ASXH 3511 beneficiary-name validation.
        c.execute("""
        CREATE TABLE IF NOT EXISTS dp01_validation_rows(
            dataset_id TEXT NOT NULL,
            row_no INTEGER NOT NULL,
            value_date TEXT NOT NULL,
            branch_code TEXT DEFAULT '',
            account_ledger TEXT DEFAULT '',
            customer_code TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            currency TEXT DEFAULT '',
            current_balance REAL DEFAULT 0,
            rate REAL DEFAULT 0,
            account_number TEXT DEFAULT '',
            search_text TEXT DEFAULT '',
            PRIMARY KEY(dataset_id,row_no),
            FOREIGN KEY(dataset_id) REFERENCES datasets(id) ON DELETE CASCADE
        )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_dp01_validation_account ON dp01_validation_rows(account_number)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_dp01_validation_date ON dp01_validation_rows(value_date)")
        # Migrate legacy mixed DP01 data once: preserve 211108 names for ASXH,
        # then remove every 211108 row from CASA balances/averages/dashboards.
        c.execute("""
            INSERT OR IGNORE INTO dp01_validation_rows
            SELECT dataset_id,row_no,value_date,branch_code,account_ledger,customer_code,
                   customer_name,product_name,currency,current_balance,rate,account_number,search_text
            FROM account_balance_rows WHERE account_ledger='211108'
        """)
        c.execute("DELETE FROM account_balance_rows WHERE account_ledger<>'421101'")
        c.execute("""
            DELETE FROM rows
            WHERE dataset_id IN (SELECT id FROM datasets WHERE dataset_kind='balance')
              AND json_extract(data_json, '$.TAI_KHOAN_HACH_TOAN')='211108'
        """)
        c.execute("""
            UPDATE datasets
            SET row_count=(SELECT count(*) FROM rows WHERE rows.dataset_id=datasets.id)
            WHERE dataset_kind='balance'
        """)
        c.execute("""
        CREATE TABLE IF NOT EXISTS precomputed_cache(
            cache_key TEXT PRIMARY KEY,
            response_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """)
        c.execute("DELETE FROM report_rows WHERE dataset_id NOT IN (SELECT id FROM datasets)")
        c.execute("DELETE FROM account_balance_rows WHERE dataset_id NOT IN (SELECT id FROM datasets)")
        c.execute("DELETE FROM rows WHERE dataset_id NOT IN (SELECT id FROM datasets)")

init_db()


def cache_get(cache_key: str) -> Optional[Dict[str, Any]]:
    with conn() as c:
        row = c.execute("SELECT response_json, updated_at FROM precomputed_cache WHERE cache_key=?", (cache_key,)).fetchone()
    if not row:
        return None
    data = json.loads(row["response_json"])
    if isinstance(data, dict):
        data["cached"] = True
        data["cache_updated_at"] = row["updated_at"]
    return data


def cache_set(cache_key: str, data: Dict[str, Any]) -> None:
    clean = dict(data)
    clean.pop("cached", None)
    clean.pop("cache_updated_at", None)
    try:
        with conn() as c:
            c.execute("""
                INSERT OR REPLACE INTO precomputed_cache(cache_key,response_json,updated_at)
                VALUES(?,?,?)
            """, (cache_key, json.dumps(clean, ensure_ascii=False), now_iso()))
    except sqlite3.OperationalError as e:
        # Dashboard must still return fresh data even if cache write waits on SQLite.
        if "locked" not in str(e).lower():
            raise


def cache_delete_prefix(prefix: str = "") -> None:
    with conn() as c:
        if prefix:
            c.execute("DELETE FROM precomputed_cache WHERE cache_key LIKE ?", (prefix + "%",))
        else:
            c.execute("DELETE FROM precomputed_cache")


def clean_col(v: Any, idx: int) -> str:
    s = str(v or "").strip().replace("\ufeff", "")
    return s or f"Cột {idx}"


def sniff_encoding(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1258", "cp1252", "latin1"):
        try:
            raw.decode(enc)
            return enc
        except UnicodeDecodeError:
            pass
    return "latin1"


def normalize_header_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())

def filter_records_by_locac_421101(cols: List[str], records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    locac_cols = [col for col in cols if normalize_header_key(col) == "LOCAC"]
    if not locac_cols:
        return records
    locac_col = locac_cols[0]
    return [r for r in records if str(r.get(locac_col, "")).strip() == "421101"]


def parse_csv(path: Path) -> tuple[List[str], List[Dict[str, str]]]:
    raw = path.read_bytes()
    enc = sniff_encoding(raw)
    text = raw.decode(enc, errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return [], []
    cols = [clean_col(v, i + 1) for i, v in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if not any(str(x).strip() for x in r):
            continue
        item = {cols[i]: str(r[i]).strip() if i < len(r) else "" for i in range(len(cols))}
        out.append(item)
    return cols, out


DP01_COLUMNS = ["MA_CN", "TAI_KHOAN_HACH_TOAN", "MA_KH", "TEN_KH", "DP_TYPE_NAME", "CCY", "CURRENT_BALANCE", "RATE", "SO_TAI_KHOAN"]


def looks_like_dp01(filename: str, cols: Optional[List[str]] = None) -> bool:
    low = (filename or "").lower()
    if "dp01" in low:
        return True
    keys = {normalize_header_key(c) for c in (cols or [])}
    return {"MACN", "TAIKHOANHACHTOAN", "MAKH", "CURRENTBALANCE", "SOTAIKHOAN"}.issubset(keys)


def looks_like_gl02(filename: str, cols: Optional[List[str]] = None) -> bool:
    low = (filename or "").lower()
    if "gl02" in low:
        return True
    keys = {normalize_header_key(c) for c in (cols or [])}
    return {"TRDATE", "CUSTOMER", "DRAMOUNT", "CRAMOUNT"}.issubset(keys) or "LOCAC" in keys


def dataset_kind_from_file(filename: str, cols: List[str]) -> str:
    if looks_like_dp01(filename, cols):
        return "balance"
    if looks_like_gl02(filename, cols):
        return "transaction"
    raise HTTPException(400, "Không nhận diện được loại file. Tên/cấu trúc phải là GL02 sao kê giao dịch hoặc DP01 sao kê số dư tài khoản.")


def parse_dp01_whitespace(path: Path) -> tuple[List[str], List[Dict[str, str]]]:
    raw = path.read_bytes()
    enc = sniff_encoding(raw)
    lines = [ln.strip() for ln in raw.decode(enc, errors="replace").splitlines() if ln.strip()]
    if not lines:
        return DP01_COLUMNS[:], []
    start = 1 if all(k in normalize_header_key(lines[0]) for k in ["MACN", "TAIKHOANHACHTOAN", "CURRENTBALANCE"]) else 0
    records: List[Dict[str, str]] = []
    product_markers = ["Tiết kiệm", "TG KKH", "Tiền gửi", "TK", "TGTK", "TG "]
    for line in lines[start:]:
        parts = line.split()
        if len(parts) < 9:
            continue
        ma_cn, ledger, ma_kh = parts[0], parts[1], parts[2].strip("'")
        so_tai_khoan = parts[-1].strip("'")
        rate = parts[-2]
        current_balance = parts[-3]
        ccy = parts[-4]
        middle = " ".join(parts[3:-4]).strip()
        ten_kh, product = middle, ""
        for marker_text in product_markers:
            pos = middle.find(marker_text)
            if pos > 0:
                ten_kh = middle[:pos].strip()
                product = middle[pos:].strip()
                break
        records.append({
            "MA_CN": ma_cn,
            "TAI_KHOAN_HACH_TOAN": ledger,
            "MA_KH": ma_kh,
            "TEN_KH": ten_kh,
            "DP_TYPE_NAME": product,
            "CCY": ccy,
            "CURRENT_BALANCE": current_balance,
            "RATE": rate,
            "SO_TAI_KHOAN": so_tai_khoan,
        })
    return DP01_COLUMNS[:], records


def parse_xlsx(path: Path) -> tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return [], []
    cols = [clean_col(v, i + 1) for i, v in enumerate(header)]
    out = []
    for row in it:
        vals = list(row or [])
        if not any(v not in (None, "") for v in vals):
            continue
        item = {cols[i]: ("" if i >= len(vals) or vals[i] is None else str(vals[i]).strip()) for i in range(len(cols))}
        out.append(item)
    return cols, out


def parse_xls_via_libreoffice(path: Path) -> tuple[List[str], List[Dict[str, str]]]:
    with tempfile.TemporaryDirectory() as td:
        outdir = Path(td)
        try:
            subprocess.run([
                "libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", str(outdir), str(path)
            ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=120)
        except FileNotFoundError:
            raise HTTPException(400, "Máy chủ chưa có LibreOffice để đọc file .xls")
        except subprocess.CalledProcessError as e:
            err = (e.stderr or b"").decode("utf-8", errors="replace")[-500:]
            raise HTTPException(400, f"Không đọc được file .xls: {err}")
        converted = next(outdir.glob("*.xlsx"), None)
        if not converted:
            raise HTTPException(400, "Không chuyển được file .xls sang .xlsx")
        return parse_xlsx(converted)


def parse_file(path: Path, filename: str):
    low = filename.lower()
    if low.endswith(".csv") or low.endswith(".txt"):
        parsed = parse_csv(path)
        cols, records = parsed
        # DP01 thực tế có thể là CSV dấu phẩy chuẩn. Chỉ dùng parser khoảng trắng khi Sniffer đọc ra 1 cột.
        if len(cols) == 1 and looks_like_dp01(filename, cols):
            return parse_dp01_whitespace(path), "csv"
        return parsed, "csv"
    if low.endswith(".xlsx") or low.endswith(".xlsm"):
        return parse_xlsx(path), "xlsx"
    if low.endswith(".xls"):
        return parse_xls_via_libreoffice(path), "xls"
    raise HTTPException(400, "Chỉ hỗ trợ CSV/TXT/XLS/XLSX")


def value_date_from_filename(filename: str) -> str:
    m = re.search(r"(20\d{6})", filename or "")
    if not m:
        return ""
    try:
        return datetime.strptime(m.group(1), "%Y%m%d").strftime("%Y-%m-%d")
    except Exception:
        return ""


def col_by_key(cols: List[str], key: str) -> Optional[str]:
    target = normalize_header_key(key)
    for c in cols:
        if normalize_header_key(c) == target:
            return c
    return None


def normalize_dp01_text(value: Any) -> str:
    """Remove whitespace and Excel/CSV text marker quotes from DP01 fields."""
    return str(value or "").strip().strip("'").strip()


def filter_dp01_rows_for_ledger(cols: List[str], records: List[Dict[str, str]], ledger: str) -> List[Dict[str, str]]:
    """Keep one DP01 ledger. Missing ledger column stays backward compatible."""
    c = col_by_key(cols, "TAI_KHOAN_HACH_TOAN")
    if not c:
        return records
    return [r for r in records if normalize_dp01_text(r.get(c, "")) == ledger]


def filter_dp01_calculation_rows(cols: List[str], records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """CASA totals, balances and averages use 421101 only."""
    return filter_dp01_rows_for_ledger(cols, records, "421101")


def filter_dp01_validation_rows(cols: List[str], records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """211108 stays available only to ASXH's 3511 name-validation lookup."""
    return filter_dp01_rows_for_ledger(cols, records, "211108")


def balance_tuple_from_record(dataset_id: str, row_no: int, cols: List[str], rec: Dict[str, Any], value_date: str):
    def get(name: str) -> str:
        c = col_by_key(cols, name)
        return normalize_dp01_text(rec.get(c, "") if c else "")
    branch = get("MA_CN")
    ledger = get("TAI_KHOAN_HACH_TOAN")
    code = get("MA_KH")
    name = get("TEN_KH")
    product = get("DP_TYPE_NAME")
    currency = get("CCY").upper()
    # DP01 rule: số dư cuối ngày luôn lấy từ cột G (CURRENT_BALANCE) của sao kê ngày đó.
    # Không suy đoán từ cột khác khi tiêu đề bị lệch/đổi tên.
    current_balance_raw = str(rec.get(cols[6], "") if len(cols) > 6 else get("CURRENT_BALANCE")).strip().strip("'")
    bal = parse_vn_number(current_balance_raw)
    rate = parse_vn_number(get("RATE"))
    acc = get("SO_TAI_KHOAN")
    search = " ".join([value_date, branch, ledger, code, name, product, currency, acc]).lower()[:5000]
    return (dataset_id, row_no, value_date, branch, ledger, code, name, product, currency, bal, rate, acc, search)


def upsert_dp01_batch(c, rows, table: str):
    if rows:
        c.executemany(f"""
        INSERT OR REPLACE INTO {table}(dataset_id,row_no,value_date,branch_code,account_ledger,customer_code,customer_name,product_name,currency,current_balance,rate,account_number,search_text)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, rows)


def upsert_balance_batch(c, rows):
    upsert_dp01_batch(c, rows, "account_balance_rows")


def upsert_validation_batch(c, rows):
    upsert_dp01_batch(c, rows, "dp01_validation_rows")


def save_dataset_from_path(path: Path, filename: str, name: Optional[str] = None, keep_file: bool = True) -> Dict[str, Any]:
    if not filename:
        raise HTTPException(400, "Thiếu file")
    dataset_id = uuid.uuid4().hex[:12]
    safe_name = re.sub(r"[^\w.\-]+", "_", filename, flags=re.UNICODE)
    stored = UPLOAD_DIR / f"{dataset_id}_{safe_name}"
    if keep_file:
        shutil.move(str(path), stored)
    else:
        shutil.copyfile(path, stored)
    try:
        (parsed, file_type) = parse_file(stored, filename)
        cols, records = parsed
        if not cols:
            raise HTTPException(400, f"File {filename} không có header")
        dataset_kind = dataset_kind_from_file(filename, cols)
        is_dp01_file = dataset_kind == "balance"
        validation_records: List[Dict[str, str]] = []
        value_date = value_date_from_filename(filename)
        if is_dp01_file:
            if not value_date:
                raise HTTPException(400, "File DP01 phải có ngày giá trị trong tên file, ví dụ 3511_dp01_20260701.csv")
            calculation_records = filter_dp01_calculation_rows(cols, records)
            validation_records = filter_dp01_validation_rows(cols, records)
            # CASA app stores/displays only 421101 rows. 211108 is written to
            # separate validation storage and is invisible to CASA calculations.
            records = calculation_records
        else:
            records = filter_records_by_locac_421101(cols, records)
        title = (name or Path(filename).stem).strip()
        with conn() as c:
            c.execute("INSERT INTO datasets(id,name,original_name,stored_path,file_type,columns_json,row_count,imported_at,dataset_kind) VALUES(?,?,?,?,?,?,?,?,?)",
                      (dataset_id, title, filename, str(stored), file_type, json.dumps(cols, ensure_ascii=False), len(records), now_iso(), dataset_kind))
            batch = []
            report_batch = []
            balance_batch = []
            validation_batch = []
            if is_dp01_file:
                for i, item in enumerate(validation_records, start=1):
                    validation_batch.append(balance_tuple_from_record(dataset_id, i, cols, item, value_date))
            for i, item in enumerate(records, start=1):
                text = " ".join(str(v).lower() for v in item.values())[:5000]
                batch.append((dataset_id, i, json.dumps(item, ensure_ascii=False), text))
                if is_dp01_file:
                    if value_date:
                        balance_batch.append(balance_tuple_from_record(dataset_id, i, cols, item, value_date))
                else:
                    report_batch.append(report_tuple_from_record(dataset_id, i, cols, item))
                if len(batch) >= 1000:
                    c.executemany("INSERT INTO rows(dataset_id,row_no,data_json,search_text) VALUES(?,?,?,?)", batch)
                    upsert_report_batch(c, report_batch)
                    upsert_balance_batch(c, balance_batch)
                    batch.clear(); report_batch.clear(); balance_batch.clear()
            # Keep 211108 in separate ASXH-validation store; never add it to
            # rows/account_balance_rows used by CASA dashboard calculations.
            upsert_validation_batch(c, validation_batch)
            if batch:
                c.executemany("INSERT INTO rows(dataset_id,row_no,data_json,search_text) VALUES(?,?,?,?)", batch)
                upsert_report_batch(c, report_batch)
                upsert_balance_batch(c, balance_batch)
        return {"id": dataset_id, "name": title, "filename": filename, "columns": cols, "row_count": len(records)}
    except Exception:
        stored.unlink(missing_ok=True)
        raise


def save_dataset_from_upload(file: UploadFile, name: Optional[str] = None) -> Dict[str, Any]:
    if not file.filename:
        raise HTTPException(400, "Thiếu file")
    temp_id = uuid.uuid4().hex
    temp_path = CHUNK_DIR / f"direct_{temp_id}.part"
    with temp_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    return save_dataset_from_path(temp_path, file.filename, name, keep_file=True)


def dataset_or_404(dataset_id: str):
    with conn() as c:
        row = c.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy dữ liệu")
    return dict(row)

def parse_vn_number(value: Any) -> float:
    s = str(value or "").strip()
    s = re.sub(r"[^0-9,\.\-]", "", s.replace(",", "."))
    if not s or s in {"-", "."}:
        return 0.0
    if s.count(".") > 1:
        parts = s.split("."); s = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(s)
    except Exception:
        return 0.0

def parse_any_date(value: Any) -> Optional[datetime]:
    raw = str(value or "").strip()
    if not raw:
        return None
    s = raw[:19].replace("T", " ")
    if re.fullmatch(r"\d{8}", raw[:8]):
        try:
            return datetime.strptime(raw[:8], "%Y%m%d")
        except Exception:
            pass
    if re.fullmatch(r"\d{8}\s+\d{2}:\d{2}:\d{2}", s):
        try:
            return datetime.strptime(s, "%Y%m%d %H:%M:%S")
        except Exception:
            pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19] if "%H" in fmt else s[:10], fmt)
        except Exception:
            pass
    return None

def infer_col(cols: List[str], keywords: List[str]) -> Optional[str]:
    for kw in keywords:
        for c in cols:
            if kw in str(c).lower():
                return c
    return None


def row_turnover_amount(rec: Dict[str, Any], turn_col: Optional[str] = None) -> float:
    if "DRAMOUNT" in rec or "CRAMOUNT" in rec:
        return parse_vn_number(rec.get("DRAMOUNT")) + parse_vn_number(rec.get("CRAMOUNT"))
    return parse_vn_number(rec.get(turn_col)) if turn_col else 0.0

def row_credit_amount(rec: Dict[str, Any], turn_col: Optional[str] = None) -> float:
    if "CRAMOUNT" in rec:
        return parse_vn_number(rec.get("CRAMOUNT"))
    return parse_vn_number(rec.get(turn_col)) if turn_col else 0.0

def row_debit_amount(rec: Dict[str, Any]) -> float:
    return parse_vn_number(rec.get("DRAMOUNT")) if "DRAMOUNT" in rec else 0.0

def report_tuple_from_record(dataset_id: str, row_no: int, cols: List[str], rec: Dict[str, Any]):
    date_col = infer_col(cols, ["trdate", "crtdtm", "crtdate", "ngày", "date", "thời gian", "ngay"])
    code_col = infer_col(cols, ["customer", "mã khách", "ma khach", "cif", "mã kh"])
    name_col = infer_col(cols, ["tên khách", "ten khach", "khách hàng", "customer name", "remark"])
    turn_col = infer_col(cols, ["dramount", "cramount", "doanh số", "doanh so", "giao dịch", "phát sinh", "transaction", "turnover", "amount"])
    dt = parse_any_date(rec.get(date_col)) if date_col else None
    tr_date = dt.strftime("%Y-%m-%d") if dt else ""
    debit = row_debit_amount(rec)
    credit = row_credit_amount(rec, turn_col)
    turnover = row_turnover_amount(rec, turn_col)
    code = str(rec.get(code_col, "") if code_col else "").strip()
    name = str(rec.get(name_col, "") if name_col else "").strip()
    currency = str(rec.get("CCY", "") or rec.get("currency", "")).strip().upper()
    search = " ".join([tr_date, code, name, currency, str(rec.get("REMARK", ""))]).lower()[:5000]
    return (dataset_id, row_no, tr_date, code, name, currency, debit, credit, turnover, search)

def upsert_report_batch(c, rows):
    if rows:
        c.executemany("""
        INSERT OR REPLACE INTO report_rows(dataset_id,row_no,tr_date,customer_code,customer_name,currency,debit,credit,turnover,search_text)
        VALUES(?,?,?,?,?,?,?,?,?,?)
        """, rows)

def dataset_records(dataset_id: Optional[str] = None, per_dataset_limit: Optional[int] = None):
    with conn() as c:
        dss = [c.execute("SELECT * FROM datasets WHERE id=?", (dataset_id,)).fetchone()] if dataset_id else c.execute("SELECT * FROM datasets ORDER BY imported_at DESC LIMIT 5").fetchall()
        out = []
        for ds in dss:
            if not ds: continue
            cols = json.loads(ds["columns_json"])
            
            if per_dataset_limit:
                recs = c.execute("SELECT data_json FROM rows WHERE dataset_id=? ORDER BY row_no DESC LIMIT ?", (ds["id"], per_dataset_limit)).fetchall()
            else:
                recs = c.execute("SELECT data_json FROM rows WHERE dataset_id=? ORDER BY row_no DESC", (ds["id"],)).fetchall()
            for r in recs:
                out.append((dict(ds), cols, json.loads(r["data_json"])))
    return out

@app.get("/", response_class=HTMLResponse)
def index():
    return (PUBLIC_DIR / "index.html").read_text(encoding="utf-8")

@app.get("/api/health")
def health():
    return {"ok": True, "port": APP_PORT, "time": now_iso()}

@app.post("/api/upload")
def upload(file: UploadFile = File(...), name: Optional[str] = Form(None)):
    result = save_dataset_from_upload(file, name)
    cache = precompute_common_views()
    return {"id": result["id"], "name": result["name"], "columns": result["columns"], "row_count": result["row_count"], "cache": cache}


@app.post("/api/upload-chunk")
def upload_chunk(
    upload_id: str = Form(...),
    chunk_index: int = Form(...),
    total_chunks: int = Form(...),
    filename: str = Form(...),
    file: UploadFile = File(...),
):
    if not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", upload_id):
        raise HTTPException(400, "upload_id không hợp lệ")
    if chunk_index < 0 or total_chunks < 1 or chunk_index >= total_chunks:
        raise HTTPException(400, "Thứ tự chunk không hợp lệ")
    session_dir = CHUNK_DIR / upload_id
    session_dir.mkdir(parents=True, exist_ok=True)
    part_path = session_dir / f"{chunk_index:06d}.part"
    with part_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"ok": True, "upload_id": upload_id, "chunk_index": chunk_index, "total_chunks": total_chunks, "filename": filename}


@app.post("/api/upload-complete")
def upload_complete(payload: Dict[str, Any] = Body(...)):
    upload_id = str(payload.get("upload_id") or "")
    filename = str(payload.get("filename") or "")
    name = payload.get("name") or None
    total_chunks = int(payload.get("total_chunks") or 0)
    if not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", upload_id):
        raise HTTPException(400, "upload_id không hợp lệ")
    if not filename or total_chunks < 1:
        raise HTTPException(400, "Thiếu thông tin file")
    session_dir = CHUNK_DIR / upload_id
    if not session_dir.exists():
        raise HTTPException(400, "Không tìm thấy phiên upload")
    missing = [i for i in range(total_chunks) if not (session_dir / f"{i:06d}.part").exists()]
    if missing:
        raise HTTPException(400, f"Thiếu chunk: {missing[:10]}")
    assembled = session_dir / "assembled.bin"
    with assembled.open("wb") as out:
        for i in range(total_chunks):
            with (session_dir / f"{i:06d}.part").open("rb") as part:
                shutil.copyfileobj(part, out)
    try:
        result = save_dataset_from_path(assembled, filename, name, keep_file=True)
        cache = precompute_common_views()
        return {"id": result["id"], "name": result["name"], "columns": result["columns"], "row_count": result["row_count"], "cache": cache}
    finally:
        shutil.rmtree(session_dir, ignore_errors=True)


@app.post("/api/upload-many")
def upload_many(files: List[UploadFile] = File(...), name: Optional[str] = Form(None)):
    if not files:
        raise HTTPException(400, "Chưa chọn file")
    imported = []
    errors = []
    for idx, file in enumerate(files, start=1):
        try:
            base_name = None
            if name and len(files) == 1:
                base_name = name
            imported.append(save_dataset_from_upload(file, base_name))
        except HTTPException as e:
            errors.append({"filename": file.filename, "error": e.detail})
        except Exception as e:
            errors.append({"filename": file.filename, "error": str(e)})
    if not imported and errors:
        raise HTTPException(400, {"message": "Không import được file nào", "errors": errors})
    cache = precompute_common_views() if imported else {"built": [], "errors": [], "updated_at": now_iso()}
    return {"count": len(imported), "imported": imported, "errors": errors, "cache": cache}

@app.get("/api/datasets")
def datasets():
    with conn() as c:
        rows = c.execute("SELECT id,name,original_name,file_type,row_count,imported_at,columns_json,dataset_kind FROM datasets ORDER BY imported_at DESC").fetchall()
    return {"datasets": [{**dict(r), "columns": json.loads(r["columns_json"])} for r in rows]}

@app.delete("/api/datasets/{dataset_id}")
def delete_dataset(dataset_id: str):
    ds = dataset_or_404(dataset_id)
    with conn() as c:
        c.execute("DELETE FROM report_rows WHERE dataset_id=?", (dataset_id,))
        c.execute("DELETE FROM account_balance_rows WHERE dataset_id=?", (dataset_id,))
        c.execute("DELETE FROM rows WHERE dataset_id=?", (dataset_id,))
        c.execute("DELETE FROM datasets WHERE id=?", (dataset_id,))
    try:
        Path(ds["stored_path"]).unlink(missing_ok=True)
    except Exception:
        pass
    cache = precompute_common_views()
    return {"ok": True, "cache": cache}

@app.get("/api/rows/{dataset_id}")
def rows(dataset_id: str, q: str = "", page: int = 1, page_size: int = 100):
    ds = dataset_or_404(dataset_id)
    page = max(1, page)
    page_size = min(max(10, page_size), 500)
    where = "dataset_id=?"
    params: List[Any] = [dataset_id]
    if q.strip():
        where += " AND search_text LIKE ?"
        params.append(f"%{q.strip().lower()}%")
    with conn() as c:
        total = c.execute(f"SELECT COUNT(*) FROM rows WHERE {where}", params).fetchone()[0]
        recs = c.execute(f"SELECT row_no,data_json FROM rows WHERE {where} ORDER BY row_no LIMIT ? OFFSET ?", params + [page_size, (page-1)*page_size]).fetchall()
    return {"dataset": ds, "columns": json.loads(ds["columns_json"]), "total": total, "page": page, "page_size": page_size,
            "rows": [{"row_no": r["row_no"], **json.loads(r["data_json"])} for r in recs]}

@app.get("/api/summary/{dataset_id}")
def summary(dataset_id: str, column: Optional[str] = None):
    ds = dataset_or_404(dataset_id)
    cols = json.loads(ds["columns_json"])
    with conn() as c:
        recs = c.execute("SELECT data_json FROM rows WHERE dataset_id=?", (dataset_id,)).fetchall()
    if not column or column not in cols:
        stats = []
        for col in cols[:20]:
            non_empty = 0
            uniques = Counter()
            for r in recs:
                v = str(json.loads(r["data_json"]).get(col, "")).strip()
                if v:
                    non_empty += 1
                    if len(uniques) < 2000:
                        uniques[v] += 1
            stats.append({"column": col, "non_empty": non_empty, "empty": len(recs)-non_empty, "unique_sample": len(uniques)})
        return {"mode": "columns", "row_count": len(recs), "stats": stats, "columns": cols}
    cnt = Counter()
    for r in recs:
        v = str(json.loads(r["data_json"]).get(column, "")).strip() or "(trống)"
        cnt[v] += 1
    return {"mode": "values", "column": column, "total": len(recs), "items": [{"value": k, "count": v} for k, v in cnt.most_common(200)]}


@app.get("/api/casa-totals")
def casa_totals():
    with conn() as c:
        latest = c.execute("SELECT MAX(value_date) AS d FROM account_balance_rows").fetchone()["d"]
        if not latest:
            return {"latest_date":"", "latest_date_label":"", "total_balance":0, "avg_balance_total":0, "customer_count":0, "row_count":0}
        latest_total = c.execute("SELECT SUM(current_balance) AS s, COUNT(*) AS n, COUNT(DISTINCT customer_code) AS c FROM account_balance_rows WHERE value_date=?", (latest,)).fetchone()
        avg_total = c.execute("""
            SELECT SUM(avg_balance) AS s FROM (
                SELECT customer_code, AVG(daily_balance) AS avg_balance
                FROM (
                    SELECT customer_code, value_date, SUM(current_balance) AS daily_balance
                    FROM account_balance_rows
                    WHERE value_date<=?
                    GROUP BY customer_code, value_date
                ) d
                GROUP BY customer_code
            )
        """, (latest,)).fetchone()["s"]
    dt = datetime.strptime(latest, "%Y-%m-%d")
    return {
        "latest_date": latest,
        "latest_date_label": dt.strftime("%d/%m/%Y"),
        "total_balance": float(latest_total["s"] or 0),
        "avg_balance_total": float(avg_total or 0),
        "customer_count": int(latest_total["c"] or 0),
        "row_count": int(latest_total["n"] or 0),
    }


def dashboard_yesterday():
    # Dashboard excludes current Vietnam-local date.
    return (datetime.utcnow() + timedelta(hours=7)).date() - timedelta(days=1)


def avg_balance_period_bounds(c, period: str):
    end_limit = dashboard_yesterday().strftime("%Y-%m-%d")
    latest = c.execute("""
        SELECT MAX(value_date) AS d
        FROM account_balance_rows
        WHERE dataset_id IN (SELECT id FROM datasets) AND value_date<=?
    """, (end_limit,)).fetchone()["d"]
    if not latest:
        return None, None, ""
    dt = datetime.strptime(latest, "%Y-%m-%d").date()
    if period == "week":
        # 7 ngày gần nhất, không tính ngày hiện tại. Ví dụ 08/07/2026 => 01/07/2026-07/07/2026.
        start = dt - timedelta(days=6)
    elif period == "month":
        start = dt.replace(day=1)
    else:
        start = None
    return (start.strftime("%Y-%m-%d") if start else None), latest, dt.strftime("%d/%m/%Y")

@app.get("/api/dashboard/top-avg-balance")
def dashboard_top_avg_balance(period: str = "all", _skip_cache: bool = False):
    cache_key = f"top_avg_balance:{period}"
    if not _skip_cache:
        cached = cache_get(cache_key)
        if cached is not None:
            return cached
    with conn() as c:
        start, latest, label = avg_balance_period_bounds(c, period)
        if not latest:
            return {"latest_date":"", "latest_date_label":"", "period": period, "rows": []}
        date_where = "b.value_date<=?"
        params = [latest]
        if start:
            date_where = "b.value_date BETWEEN ? AND ?"
            params = [start, latest]
        recs = c.execute(f"""
            SELECT d.customer_code, MAX(d.customer_name) AS customer_name, d.currency,
                   COALESCE(MAX(t.officer_name), '') AS registered_officer,
                   COUNT(*) AS day_count,
                   AVG(d.daily_balance) AS avg_balance,
                   SUM(CASE WHEN d.value_date=? THEN d.daily_balance ELSE 0 END) AS total_balance
            FROM (
                SELECT b.customer_code, MAX(b.customer_name) AS customer_name, b.currency, b.value_date,
                       SUM(b.current_balance) AS daily_balance
                FROM account_balance_rows b
                WHERE b.dataset_id IN (SELECT id FROM datasets) AND {date_where}
                GROUP BY b.customer_code, b.currency, b.value_date
            ) d
            LEFT JOIN customer_targets t ON t.customer_code=d.customer_code
            GROUP BY d.customer_code, d.currency
            HAVING AVG(d.daily_balance) <> 0
            ORDER BY avg_balance DESC, d.customer_code
            LIMIT 10
        """, [latest] + params).fetchall()
    result = {
        "latest_date": latest,
        "latest_date_label": label,
        "period": period,
        "rows": [{
            "customer_code": r["customer_code"] or "(trống)",
            "customer_name": r["customer_name"] or "",
            "currency": r["currency"] or "",
            "registered_officer": r["registered_officer"] if "registered_officer" in r.keys() else "",
            "day_count": int(r["day_count"] or 0),
            "avg_balance": float(r["avg_balance"] or 0),
            "total_balance": float(r["total_balance"] or 0),
        } for r in recs]
    }
    if not _skip_cache:
        cache_set(cache_key, result)
    return result

@app.get("/api/dashboard/top-officer-avg-balance")
def dashboard_top_officer_avg_balance(period: str = "all", ranking: str = "high", _skip_cache: bool = False):
    ranking = "low" if ranking == "low" else "high"
    cache_key = f"top_officer_avg_balance:v6:{period}:{ranking}"
    if not _skip_cache:
        cached = cache_get(cache_key)
        if cached is not None:
            return cached
    # Rule: app dkdv holds TKTT hoạt động registrations by account number.
    # Example: Đỗ Doãn Lộc registers 3511201017720 => find MA_KH for that account in CASA balance rows,
    # then add that account/customer average balance to officer achievement.
    dkdv_targets = load_dkdv_tktt_targets()
    local_rows: List[Dict[str, str]] = []
    with conn() as c:
        start, latest, label = avg_balance_period_bounds(c, period)
        if not latest:
            return {"latest_date":"", "latest_date_label":"", "period": period, "source":"dkdv", "rows": []}
        date_where = "b.value_date<=?"
        date_params = [latest]
        if start:
            date_where = "b.value_date BETWEEN ? AND ?"
            date_params = [start, latest]
        if dkdv_targets:
            c.execute("CREATE TEMP TABLE IF NOT EXISTS tmp_dkdv_tktt_targets(account_number TEXT, officer_name TEXT, customer_name TEXT)")
            c.execute("DELETE FROM tmp_dkdv_tktt_targets")
            c.executemany(
                "INSERT INTO tmp_dkdv_tktt_targets(account_number, officer_name, customer_name) VALUES (?,?,?)",
                [(x["account_number"], x["officer_name"], x.get("customer_name", "")) for x in dkdv_targets]
            )
            target_table = "tmp_dkdv_tktt_targets"
            source = "dkdv"
        else:
            # Fallback: manually entered config targetCode still interpreted as account number.
            c.execute("CREATE TEMP TABLE IF NOT EXISTS tmp_dkdv_tktt_targets(account_number TEXT, officer_name TEXT, customer_name TEXT)")
            c.execute("DELETE FROM tmp_dkdv_tktt_targets")
            c.execute("""
                INSERT INTO tmp_dkdv_tktt_targets(account_number, officer_name, customer_name)
                SELECT customer_code, officer_name, customer_name FROM customer_targets
            """)
            target_table = "tmp_dkdv_tktt_targets"
            source = "local_config"
        excluded_low_officers = [
            "Nguyễn Quốc Huy", "Đỗ Văn Nam", "Nguyễn Chí Thanh", "Nguyễn Ngọc Tú",
            "Đặng Thị Hảo", "Trần Thị Hồ Lan", "Phạm Văn Khoa", "Trịnh Ngọc Nam", "Đỗ Doãn Lộc",
            "Lê Thị Giang", "Lê Thu Phương", "Nguyễn Quốc Vương Linh",
        ]
        exclusion_sql = ""
        query_params = list(date_params)
        if ranking == "low":
            exclusion_sql = "WHERE officer_name NOT IN (" + ",".join("?" for _ in excluded_low_officers) + ")"
            query_params.extend(excluded_low_officers)
        order_sql = "ASC" if ranking == "low" else "DESC"
        having_sql = "" if ranking == "low" else "WHERE total_avg_balance <> 0"
        limit_sql = "" if ranking == "low" else "LIMIT 10"
        recs = c.execute(f"""
            WITH officer_customers AS (
                SELECT DISTINCT t.officer_name, seed.customer_code
                FROM {target_table} t
                JOIN account_balance_rows seed ON seed.account_number=t.account_number
                WHERE TRIM(COALESCE(t.officer_name,''))<>''
                  AND TRIM(COALESCE(seed.customer_code,''))<>''
            ), daily_accounts AS (
                SELECT oc.officer_name, b.customer_code, b.account_number, b.value_date,
                       SUM(b.current_balance) AS daily_balance
                FROM officer_customers oc
                JOIN account_balance_rows b ON b.customer_code=oc.customer_code
                WHERE b.dataset_id IN (SELECT id FROM datasets)
                  AND {date_where}
                  AND TRIM(COALESCE(b.account_number,''))<>''
                GROUP BY oc.officer_name, b.customer_code, b.account_number, b.value_date
            ), account_averages AS (
                SELECT officer_name, customer_code, account_number, AVG(daily_balance) AS avg_balance
                FROM daily_accounts
                GROUP BY officer_name, customer_code, account_number
            ), officer_totals AS (
                SELECT officer_name,
                       COUNT(DISTINCT CASE WHEN avg_balance > 0 THEN customer_code END) AS customer_count,
                       COUNT(DISTINCT CASE WHEN avg_balance > 0 THEN account_number END) AS account_count,
                       SUM(avg_balance) AS total_avg_balance
                FROM account_averages
                GROUP BY officer_name
            )
            SELECT officer_name, customer_count, account_count, total_avg_balance
            FROM officer_totals
            {exclusion_sql}
            {having_sql}
            ORDER BY total_avg_balance {order_sql}, officer_name
            {limit_sql}
        """, query_params).fetchall()
        if ranking == "low":
            ranked_names = {str(r["officer_name"] or "").strip() for r in recs}
            placeholders = ",".join("?" for _ in excluded_low_officers)
            zero_officers = c.execute(f"""
                SELECT DISTINCT officer_name FROM (
                    SELECT TRIM(officer_name) AS officer_name FROM {target_table}
                    UNION
                    SELECT TRIM(full_name) AS officer_name FROM app_users
                ) officers
                WHERE TRIM(COALESCE(officer_name,''))<>''
                  AND TRIM(officer_name) NOT IN ({placeholders})
                ORDER BY officer_name
            """, excluded_low_officers).fetchall()
            recs = list(recs) + [
                {"officer_name": r["officer_name"], "customer_count": 0, "account_count": 0, "total_avg_balance": 0}
                for r in zero_officers if r["officer_name"] not in ranked_names
            ]
            recs = sorted(recs, key=lambda r: (float(r["total_avg_balance"] or 0), str(r["officer_name"] or "")))[:10]
        if source == "dkdv":
            local_rows = [dict(r) for r in c.execute(f"""
                SELECT officer_name, account_number, customer_code, MAX(customer_name) AS customer_name,
                       AVG(daily_balance) AS avg_balance
                FROM (
                    SELECT t.officer_name, t.account_number, b.customer_code, MAX(b.customer_name) AS customer_name,
                           b.value_date, SUM(b.current_balance) AS daily_balance
                    FROM tmp_dkdv_tktt_targets t
                    JOIN account_balance_rows b ON t.account_number=b.account_number
                    WHERE b.dataset_id IN (SELECT id FROM datasets) AND {date_where}
                    GROUP BY t.officer_name, t.account_number, b.customer_code, b.value_date
                ) daily_match
                GROUP BY officer_name, account_number, customer_code
                ORDER BY officer_name, account_number
                LIMIT 500
            """, date_params).fetchall()]
    result = {
        "latest_date": latest,
        "latest_date_label": label,
        "period": period,
        "ranking": ranking,
        "source": source,
        "dkdv_count": len(dkdv_targets),
        "matched_accounts": len(local_rows),
        "rows": [{
            "officer_name": r["officer_name"] or "",
            "customer_count": int(r["customer_count"] or 0),
            "account_count": int(r["account_count"] or 0),
            "total_avg_balance": float(r["total_avg_balance"] or 0),
        } for r in recs],
        "matches": local_rows,
    }
    if not _skip_cache:
        cache_set(cache_key, result)
    return result

@app.get("/api/dashboard/top-group-avg-balance")
def dashboard_top_group_avg_balance(period: str = "all", _skip_cache: bool = False):
    cache_key = f"top_group_avg_balance:{period}"
    if not _skip_cache:
        cached = cache_get(cache_key)
        if cached is not None:
            return cached
    dkdv_targets = load_dkdv_tktt_targets()
    with conn() as c:
        start, latest, label = avg_balance_period_bounds(c, period)
        if not latest:
            return {"latest_date":"", "latest_date_label":"", "period": period, "rows": []}
        date_where = "b.value_date<=?"
        date_params = [latest]
        if start:
            date_where = "b.value_date BETWEEN ? AND ?"
            date_params = [start, latest]
        c.execute("CREATE TEMP TABLE IF NOT EXISTS tmp_dkdv_tktt_targets(account_number TEXT, officer_name TEXT, customer_name TEXT)")
        c.execute("DELETE FROM tmp_dkdv_tktt_targets")
        if dkdv_targets:
            c.executemany(
                "INSERT INTO tmp_dkdv_tktt_targets(account_number, officer_name, customer_name) VALUES (?,?,?)",
                [(x["account_number"], x["officer_name"], x.get("customer_name", "")) for x in dkdv_targets]
            )
        else:
            c.execute("""
                INSERT INTO tmp_dkdv_tktt_targets(account_number, officer_name, customer_name)
                SELECT customer_code, officer_name, customer_name FROM customer_targets
            """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_tmp_dkdv_tktt_account ON tmp_dkdv_tktt_targets(account_number)")
        recs = c.execute(f"""
            SELECT g.name AS group_name,
                   COUNT(DISTINCT x.officer_name) AS officer_count,
                   COUNT(DISTINCT x.customer_code) AS customer_count,
                   COUNT(DISTINCT x.account_number) AS account_count,
                   SUM(x.avg_balance) AS total_avg_balance
            FROM (
                SELECT officer_name, customer_code, account_number, AVG(daily_balance) AS avg_balance
                FROM (
                    SELECT t.officer_name AS officer_name,
                           b.customer_code AS customer_code,
                           b.account_number AS account_number,
                           b.value_date AS value_date,
                           SUM(b.current_balance) AS daily_balance
                    FROM account_balance_rows b
                    JOIN tmp_dkdv_tktt_targets t ON t.account_number=b.account_number
                    WHERE b.dataset_id IN (SELECT id FROM datasets)
                      AND {date_where}
                      AND TRIM(COALESCE(t.officer_name,''))<>''
                      AND TRIM(COALESCE(b.account_number,''))<>''
                    GROUP BY t.officer_name, b.customer_code, b.account_number, b.value_date
                ) daily_account
                GROUP BY officer_name, customer_code, account_number
            ) x
            JOIN app_users u ON LOWER(TRIM(u.full_name))=LOWER(TRIM(x.officer_name))
            JOIN user_group_members gm ON gm.user_id=u.id
            JOIN user_groups g ON g.id=gm.group_id
            GROUP BY g.id, g.name
            HAVING SUM(x.avg_balance) <> 0
            ORDER BY total_avg_balance DESC, g.name
        """, date_params).fetchall()
    result = {
        "latest_date": latest,
        "latest_date_label": label,
        "period": period,
        "rows": [{
            "group_name": r["group_name"] or "",
            "officer_count": int(r["officer_count"] or 0),
            "customer_count": int(r["customer_count"] or 0),
            "account_count": int(r["account_count"] or 0),
            "total_avg_balance": float(r["total_avg_balance"] or 0),
        } for r in recs]
    }
    if not _skip_cache:
        cache_set(cache_key, result)
    return result


@app.get("/api/dashboard")
def dashboard(dataset_id: Optional[str] = None, _skip_cache: bool = False):
    cache_key = "dashboard:all" if not dataset_id else ""
    if cache_key and not _skip_cache:
        cached = cache_get(cache_key)
        if cached is not None:
            return cached
    where = "r.tr_date != '' AND r.dataset_id IN (SELECT id FROM datasets)"
    params: List[Any] = []
    if dataset_id:
        where += " AND r.dataset_id=?"
        params.append(dataset_id)
    with conn() as c:
        end_limit = dashboard_yesterday().strftime("%Y-%m-%d")
        max_date = c.execute(f"SELECT MAX(r.tr_date) AS d FROM report_rows r WHERE {where} AND r.tr_date<=?", params + [end_limit]).fetchone()["d"]
        max_balance_date = c.execute("""
            SELECT MAX(b.value_date) AS d
            FROM account_balance_rows b
            WHERE b.dataset_id IN (SELECT id FROM datasets) AND b.value_date<=?
        """, (end_limit,)).fetchone()["d"]
        date_candidates = [d for d in [max_date, max_balance_date] if d]
        if date_candidates:
            today = datetime.strptime(max(date_candidates), "%Y-%m-%d").date()
        else:
            today = dashboard_yesterday()
        days = [today - timedelta(days=i) for i in range(6, -1, -1)]
        day_keys = [d.strftime("%Y-%m-%d") for d in days]
        qmarks = ",".join(["?"] * len(day_keys))
        q_where = f"r.tr_date IN ({qmarks}) AND r.dataset_id IN (SELECT id FROM datasets)"
        q_params: List[Any] = day_keys[:]
        if dataset_id:
            q_where += " AND r.dataset_id=?"
            q_params.append(dataset_id)
        rows = c.execute(f"SELECT r.tr_date, COUNT(*) AS cnt, SUM(r.turnover) AS turnover FROM report_rows r WHERE {q_where} GROUP BY r.tr_date", q_params).fetchall()
        b_where = f"b.value_date IN ({qmarks}) AND b.dataset_id IN (SELECT id FROM datasets)"
        b_params: List[Any] = day_keys[:]
        if dataset_id:
            b_where += " AND b.dataset_id=?"
            b_params.append(dataset_id)
        b_rows = c.execute(f"SELECT b.value_date, COUNT(*) AS cnt, SUM(b.current_balance) AS balance FROM account_balance_rows b WHERE {b_where} GROUP BY b.value_date", b_params).fetchall()
    m = {r["tr_date"]: r for r in rows}
    bm = {r["value_date"]: r for r in b_rows}
    result = {
        "labels": [d.strftime("%d/%m") for d in days],
        "balance": [round(float(bm.get(k, {"balance": 0})["balance"] or 0), 2) for k in day_keys],
        "turnover": [round(float(m.get(k, {"turnover": 0})["turnover"] or 0), 2) for k in day_keys],
        "row_count": int(sum(int(m.get(k, {"cnt": 0})["cnt"] or 0) for k in day_keys)) + int(sum(int(bm.get(k, {"cnt": 0})["cnt"] or 0) for k in day_keys)),
    }
    if cache_key and not _skip_cache:
        cache_set(cache_key, result)
    return result

@app.get("/api/report/turnover")
def report_turnover(dataset_id: Optional[str] = None, from_date: str = "", to_date: str = "", q: str = "", currency: str = "", officer: str = ""):
    where = "r.dataset_id IN (SELECT id FROM datasets)"
    params: List[Any] = []
    if dataset_id:
        where += " AND r.dataset_id=?"
        params.append(dataset_id)
    f = parse_any_date(from_date).strftime("%Y-%m-%d") if from_date and parse_any_date(from_date) else ""
    t = parse_any_date(to_date).strftime("%Y-%m-%d") if to_date and parse_any_date(to_date) else ""
    if f:
        where += " AND r.tr_date>=?"
        params.append(f)
    if t:
        where += " AND r.tr_date<=?"
        params.append(t)
    if q.strip():
        where += " AND r.search_text LIKE ?"
        params.append(f"%{q.strip().lower()}%")
    cur = str(currency or "").strip().upper()
    if cur:
        where += " AND UPPER(r.currency)=?"
        params.append(cur)
    officer_name = str(officer or "").strip()
    if officer_name:
        # SQLite LOWER() is ASCII-only; Vietnamese Đ/đ breaks LOWER(t.officer_name) LIKE lower(param).
        where += " AND EXISTS (SELECT 1 FROM customer_targets t WHERE t.customer_code=r.customer_code AND t.officer_name LIKE ? COLLATE NOCASE)"
        params.append(f"%{officer_name}%")
    with conn() as c:
        meta = c.execute(f"SELECT SUM(r.credit) AS total_credit FROM report_rows r WHERE {where}", params).fetchone()
        recs = c.execute(f"""
            SELECT r.customer_code AS customer_code, r.currency AS currency, COUNT(*) AS transaction_count, SUM(r.credit) AS total_credit
            FROM report_rows r
            WHERE {where}
            GROUP BY r.currency, r.customer_code
            HAVING SUM(r.credit) <> 0
            ORDER BY CASE UPPER(COALESCE(r.currency,'')) WHEN 'VND' THEN 1 WHEN 'USD' THEN 2 WHEN 'EUR' THEN 3 ELSE 4 END,
                     UPPER(COALESCE(r.currency,'')), total_credit DESC, transaction_count DESC, r.customer_code

        """, params).fetchall()
    rows = []
    for r in recs:
        rows.append({
            "customer_code": r["customer_code"] or "(trống)",
            "currency": r["currency"] or "(trống)",
            "transaction_count": int(r["transaction_count"] or 0),
            "total_credit": float(r["total_credit"] or 0),
        })
    report_total = sum(float(r.get("total_credit") or 0) for r in rows)
    return {"total": round(report_total, 2), "count": len(rows), "limited": False, "rows": rows}

@app.get("/api/report/balance")
def report_balance(from_date: str = "", to_date: str = "", q: str = "", currency: str = "", officer: str = "", _skip_cache: bool = False):
    use_cache = not any(str(x or "").strip() for x in [from_date, to_date, q, currency, officer])
    cache_key = "report_balance:default" if use_cache else ""
    if cache_key and not _skip_cache:
        cached = cache_get(cache_key)
        if cached is not None:
            return cached
    where = "b.dataset_id IN (SELECT id FROM datasets)"
    params: List[Any] = []
    f = parse_any_date(from_date).strftime("%Y-%m-%d") if from_date and parse_any_date(from_date) else ""
    t = parse_any_date(to_date).strftime("%Y-%m-%d") if to_date and parse_any_date(to_date) else ""
    if f:
        where += " AND b.value_date>=?"; params.append(f)
    if t:
        where += " AND b.value_date<=?"; params.append(t)
    if q.strip():
        where += " AND b.search_text LIKE ?"; params.append(f"%{q.strip().lower()}%")
    cur = str(currency or "").strip().upper()
    if cur:
        where += " AND UPPER(b.currency)=?"; params.append(cur)
    officer_name = str(officer or "").strip()
    if officer_name:
        # SQLite LOWER() is ASCII-only; Vietnamese Đ/đ breaks LOWER(t.officer_name) LIKE lower(param).
        where += " AND EXISTS (SELECT 1 FROM customer_targets t WHERE t.customer_code=b.customer_code AND t.officer_name LIKE ? COLLATE NOCASE)"; params.append(f"%{officer_name}%")
    with conn() as c:
        latest_in_period = c.execute(f"SELECT MAX(b.value_date) AS d FROM account_balance_rows b WHERE {where}", params).fetchone()["d"] or ""
        daily = c.execute(f"""
            SELECT b.value_date, b.currency, COUNT(*) AS account_count, SUM(b.current_balance) AS total_balance
            FROM account_balance_rows b WHERE {where}
            GROUP BY b.value_date, b.currency
            ORDER BY b.value_date, CASE UPPER(COALESCE(b.currency,'')) WHEN 'VND' THEN 1 WHEN 'USD' THEN 2 WHEN 'EUR' THEN 3 ELSE 4 END
        """, params).fetchall()
        avg = c.execute(f"""
            SELECT d.customer_code, MAX(d.customer_name) AS customer_name, d.currency,
                   COALESCE(MAX(t.officer_name), '') AS registered_officer,
                   COUNT(*) AS day_count,
                   AVG(d.daily_balance) AS avg_balance,
                   SUM(CASE WHEN d.value_date=? THEN d.daily_balance ELSE 0 END) AS total_balance
            FROM (
                SELECT b.customer_code, MAX(b.customer_name) AS customer_name, b.currency, b.value_date,
                       SUM(b.current_balance) AS daily_balance
                FROM account_balance_rows b WHERE {where}
                GROUP BY b.customer_code, b.currency, b.value_date
            ) d
            LEFT JOIN customer_targets t ON t.customer_code=d.customer_code
            GROUP BY d.customer_code, d.currency
            HAVING AVG(d.daily_balance) <> 0
            ORDER BY CASE UPPER(COALESCE(d.currency,'')) WHEN 'VND' THEN 1 WHEN 'USD' THEN 2 WHEN 'EUR' THEN 3 ELSE 4 END, avg_balance DESC, d.customer_code

        """, [latest_in_period] + params).fetchall()
    daily_rows = [{"value_date": r["value_date"], "currency": r["currency"] or "", "account_count": int(r["account_count"] or 0), "total_balance": float(r["total_balance"] or 0)} for r in daily]
    avg_rows = [{"customer_code": r["customer_code"] or "(trống)", "customer_name": r["customer_name"] or "", "currency": r["currency"] or "", "registered_officer": r["registered_officer"] or "", "day_count": int(r["day_count"] or 0), "avg_balance": float(r["avg_balance"] or 0), "total_balance": float(r["total_balance"] or 0)} for r in avg]
    result = {"daily": daily_rows, "rows": avg_rows, "count": len(avg_rows), "limited": False, "total_daily_balance": sum(r["total_balance"] for r in daily_rows), "total_avg_balance": sum(r["avg_balance"] for r in avg_rows)}
    if cache_key and not _skip_cache:
        cache_set(cache_key, result)
    return result

@app.get("/api/report/customer-balance")
def report_customer_balance(customer_code: str = "", from_date: str = "", to_date: str = ""):
    code = str(customer_code or "").strip()
    if not code:
        raise HTTPException(400, "Nhập mã KH")
    where = "b.dataset_id IN (SELECT id FROM datasets) AND b.customer_code=?"
    params: List[Any] = [code]
    f = parse_any_date(from_date).strftime("%Y-%m-%d") if from_date and parse_any_date(from_date) else ""
    t = parse_any_date(to_date).strftime("%Y-%m-%d") if to_date and parse_any_date(to_date) else ""
    if f:
        where += " AND b.value_date>=?"; params.append(f)
    if t:
        where += " AND b.value_date<=?"; params.append(t)
    with conn() as c:
        recs = c.execute(f"""
            SELECT b.value_date, MAX(b.customer_name) AS customer_name,
                   GROUP_CONCAT(DISTINCT b.account_number) AS account_numbers,
                   COALESCE(MAX(t.officer_name), '') AS registered_officer,
                   SUM(b.current_balance) AS casa_balance
            FROM account_balance_rows b
            LEFT JOIN customer_targets t ON t.customer_code=b.customer_code
            WHERE {where}
            GROUP BY b.value_date
            ORDER BY b.value_date
        """, params).fetchall()
    rows = []
    running_total = 0.0
    for i, r in enumerate(recs, start=1):
        bal = float(r["casa_balance"] or 0)
        running_total += bal
        rows.append({
            "row_no": i,
            "value_date": datetime.strptime(r["value_date"], "%Y-%m-%d").strftime("%d/%m/%Y") if r["value_date"] else "",
            "value_date_raw": r["value_date"] or "",
            "account_number": r["account_numbers"] or "",
            "registered_officer": r["registered_officer"] or "",
            "casa_balance": bal,
            "avg_balance": running_total / i,
            "customer_name": r["customer_name"] or "",
        })
    return {"customer_code": code, "customer_name": rows[-1]["customer_name"] if rows else "", "registered_officer": rows[-1]["registered_officer"] if rows else "", "count": len(rows), "rows": rows}

@app.get("/api/report/turnover.xlsx")
def report_turnover_xlsx(from_date: str = "", to_date: str = "", q: str = "", currency: str = "", officer: str = ""):
    data = report_turnover(None, from_date, to_date, q, currency, officer)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao doanh so"
    ws.append(["STT", "Mã KH", "Loại tiền tệ", "Số lượng giao dịch", "Tổng doanh số"] )
    for i, r in enumerate(data["rows"], start=1):
        ws.append([i, r.get("customer_code", ""), r.get("currency", ""), r.get("transaction_count", 0), r.get("total_credit", 0)])
    ws.append([])
    ws.append(["Tổng số dòng báo cáo", data["count"]])
    ws.append(["Tổng doanh số báo có", data["total"]])
    out = EXPORT_DIR / f"bao-cao-doanh-so-theo-ma-kh-{int(datetime.now().timestamp())}.xlsx"
    wb.save(out)
    return FileResponse(out, filename="bao-cao-doanh-so-theo-ma-kh.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/config")
def config_get():
    with conn() as c:
        return {"users":[dict(r) for r in c.execute("SELECT * FROM app_users ORDER BY username")], "groups":[dict(r) for r in c.execute("SELECT * FROM user_groups ORDER BY name")], "members":[dict(r) for r in c.execute("SELECT * FROM user_group_members")], "targets":[dict(r) for r in c.execute("SELECT * FROM customer_targets ORDER BY customer_code")]}

def upsert(table, fields, payload):
    item_id = str(payload.get("id") or uuid.uuid4().hex[:12]); vals=[str(payload.get(f) or "").strip() for f in fields]
    with conn() as c:
        qs=','.join(['?']*(len(fields)+2)); cols='id,'+','.join(fields)+',created_at'
        c.execute(f"INSERT OR REPLACE INTO {table}({cols}) VALUES({qs})", [item_id]+vals+[now_iso()])
    return {"ok": True, "id": item_id}

@app.post("/api/config/users")
def save_user(payload: Dict[str, Any] = Body(...)):
    if not payload.get('username') or not payload.get('full_name'): raise HTTPException(400,'Thiếu user hoặc tên người dùng')
    result = upsert('app_users',['username','full_name','department','unit'],payload)
    group_id = str(payload.get('group_id') or '').strip()
    with conn() as c:
        c.execute('DELETE FROM user_group_members WHERE user_id=?',(result['id'],))
        if group_id:
            c.execute('INSERT OR IGNORE INTO user_group_members(group_id,user_id) VALUES(?,?)',(group_id,result['id']))
    return result
@app.delete("/api/config/users/{item_id}")
def del_user(item_id: str):
    with conn() as c: c.execute('DELETE FROM user_group_members WHERE user_id=?',(item_id,)); c.execute('DELETE FROM app_users WHERE id=?',(item_id,))
    return {"ok": True}
@app.post("/api/config/groups")
def save_group(payload: Dict[str, Any] = Body(...)):
    if not payload.get('name'): raise HTTPException(400,'Thiếu tên nhóm')
    res=upsert('user_groups',['name','note'],payload); members=[str(x) for x in payload.get('members',[])]
    with conn() as c:
        c.execute('DELETE FROM user_group_members WHERE group_id=?',(res['id'],)); c.executemany('INSERT OR IGNORE INTO user_group_members(group_id,user_id) VALUES(?,?)',[(res['id'],m) for m in members])
    return res
@app.delete("/api/config/groups/{item_id}")
def del_group(item_id: str):
    with conn() as c: c.execute('DELETE FROM user_group_members WHERE group_id=?',(item_id,)); c.execute('DELETE FROM user_groups WHERE id=?',(item_id,))
    return {"ok": True}
@app.post("/api/config/targets")
def save_target(payload: Dict[str, Any] = Body(...)):
    if not payload.get('customer_code') or not payload.get('officer_name'): raise HTTPException(400,'Thiếu mã khách hàng hoặc cán bộ')
    return upsert('customer_targets',['customer_code','customer_name','officer_name','note'],payload)

@app.post("/api/sync/dkdv-tktt")
def sync_dkdv_tktt(payload: Dict[str, Any] = Body(...)):
    account = re.sub(r"\D", "", str(payload.get("account_number") or payload.get("accountNumber") or payload.get("Số tài khoản") or ""))
    officer = str(payload.get("officer_name") or payload.get("staffName") or payload.get("Tên cán bộ") or "").strip()
    customer_name = str(payload.get("customer_name") or payload.get("customerName") or payload.get("Tên khách hàng") or "").strip()
    if not account or not officer:
        raise HTTPException(400, "Thiếu số tài khoản hoặc cán bộ đăng ký")
    with conn() as c:
        row = c.execute("""
            SELECT customer_code, COALESCE(NULLIF(customer_name,''), ?) AS customer_name
            FROM account_balance_rows
            WHERE REPLACE(REPLACE(REPLACE(COALESCE(account_number,''),' ',''),'-',''),'.','')=?
            ORDER BY value_date DESC, row_no DESC
            LIMIT 1
        """, (customer_name, account)).fetchone()
        if not row or not (row["customer_code"] or "").strip():
            return {"ok": False, "synced": False, "reason": "Chưa tìm thấy mã KH trong dữ liệu CASA 8099", "account_number": account}
        code = str(row["customer_code"]).strip()
        name = str(row["customer_name"] or customer_name).strip()
        old = c.execute("SELECT id, note FROM customer_targets WHERE customer_code=? LIMIT 1", (code,)).fetchone()
        item_id = old["id"] if old else f"dkdv_{code}"
        note = old["note"] if old else "Đồng bộ tự động từ app ĐKDV 8091"
        c.execute("""
            INSERT OR REPLACE INTO customer_targets(id, customer_code, customer_name, officer_name, note, created_at)
            VALUES(?,?,?,?,?,?)
        """, (item_id, code, name, officer, note, now_iso()))
    return {"ok": True, "synced": True, "account_number": account, "customer_code": code, "customer_name": name, "officer_name": officer}
@app.delete("/api/config/targets/{item_id}")
def del_target(item_id: str):
    with conn() as c: c.execute('DELETE FROM customer_targets WHERE id=?',(item_id,))
    return {"ok": True}

def precompute_common_views() -> Dict[str, Any]:
    cache_delete_prefix("")
    built = []
    errors = []
    jobs = [
        ("dashboard:all", lambda: dashboard(_skip_cache=True)),
        ("report_balance:default", lambda: report_balance(_skip_cache=True)),
    ]
    for period in ("all", "month", "week"):
        jobs.extend([
            (f"top_avg_balance:{period}", lambda p=period: dashboard_top_avg_balance(period=p, _skip_cache=True)),
            (f"top_officer_avg_balance:{period}", lambda p=period: dashboard_top_officer_avg_balance(period=p, _skip_cache=True)),
            (f"top_group_avg_balance:{period}", lambda p=period: dashboard_top_group_avg_balance(period=p, _skip_cache=True)),
        ])
    for key, fn in jobs:
        try:
            data = fn()
            cache_set(key, data)
            built.append(key)
        except Exception as e:
            errors.append({"key": key, "error": str(e)})
    return {"built": built, "errors": errors, "updated_at": now_iso()}


@app.post("/api/cache/rebuild")
def rebuild_cache():
    return precompute_common_views()


@app.get("/api/export/{dataset_id}.xlsx")
def export_xlsx(dataset_id: str, q: str = ""):
    ds = dataset_or_404(dataset_id)
    cols = json.loads(ds["columns_json"])
    where = "dataset_id=?"
    params: List[Any] = [dataset_id]
    if q.strip():
        where += " AND search_text LIKE ?"
        params.append(f"%{q.strip().lower()}%")
    with conn() as c:
        recs = c.execute(f"SELECT row_no,data_json FROM rows WHERE {where} ORDER BY row_no", params).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Du lieu"
    ws.append(["STT"] + cols)
    for r in recs:
        d = json.loads(r["data_json"])
        ws.append([r["row_no"]] + [d.get(col, "") for col in cols])
    out = EXPORT_DIR / f"{dataset_id}-{int(datetime.now().timestamp())}.xlsx"
    wb.save(out)
    return FileResponse(out, filename=f"{ds['name']}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/export/{dataset_id}.pdf")
def export_pdf(dataset_id: str, q: str = ""):
    ds = dataset_or_404(dataset_id)
    cols = json.loads(ds["columns_json"])[:8]
    where = "dataset_id=?"
    params: List[Any] = [dataset_id]
    if q.strip():
        where += " AND search_text LIKE ?"
        params.append(f"%{q.strip().lower()}%")
    with conn() as c:
        recs = c.execute(f"SELECT row_no,data_json FROM rows WHERE {where} ORDER BY row_no LIMIT 1000", params).fetchall()
    out = EXPORT_DIR / f"{dataset_id}-{int(datetime.now().timestamp())}.pdf"
    doc = SimpleDocTemplate(str(out), pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('PdfTitleTNR', parent=styles['Title'], fontName='TimesNewRoman-Bold')
    normal_style = ParagraphStyle('PdfNormalTNR', parent=styles['Normal'], fontName='TimesNewRoman')
    body = [Paragraph(f"BÁO CÁO DỮ LIỆU: {ds['name']}", title_style), Paragraph(f"Số dòng xuất PDF: {len(recs)} (tối đa 1000 dòng)", normal_style), Spacer(1, 8)]
    table_data = [["STT"] + cols]
    for r in recs:
        d = json.loads(r["data_json"])
        table_data.append([str(r["row_no"])] + [str(d.get(col, ""))[:60] for col in cols])
    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'TimesNewRoman'), ('FONTNAME', (0,0), (-1,0), 'TimesNewRoman-Bold'), ('FONTSIZE', (0,0), (-1,-1), 7),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey), ('VALIGN', (0,0), (-1,-1), 'TOP')
    ]))
    body.append(tbl)
    doc.build(body)
    return FileResponse(out, filename=f"{ds['name']}.pdf", media_type="application/pdf")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=APP_PORT)
